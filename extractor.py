import openpyxl


def _normalize_header(text):
    """去除表头中的换行符和空白，用于匹配。"""
    if text is None:
        return ""
    return str(text).replace("\n", "").replace("\r", "").strip()


def _build_column_map(header_row, columns_config):
    """
    根据表头行构建列索引映射。

    参数:
        header_row: openpyxl 行对象
        columns_config: {目标字段: 表头关键词}

    返回:
        {列号(0-based): 目标字段名}
    """
    col_map = {}
    for idx, cell in enumerate(header_row):
        header_text = _normalize_header(cell.value)
        for target_field, keyword in columns_config.items():
            if header_text.startswith(keyword):
                col_map[idx] = target_field
                break
    return col_map


def extract_file(filepath, config):
    """
    从 Excel 文件提取目标字段数据。

    参数:
        filepath: Path 对象或字符串
        config: 该文件的配置 dict (来自 file_configs.py)

    Yields:
        dict: {目标字段: 原始值, "_source_row": 行号}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    header_row = list(sheet.iter_rows(
        min_row=config["header_row"],
        max_row=config["header_row"]
    ))[0]

    col_map = _build_column_map(header_row, config["columns"])

    for row_idx in range(config["data_start_row"], sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]

        # 跳过全空行
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue

        record = {"_source_row": row_idx}
        for col_idx, target_field in col_map.items():
            record[target_field] = row[col_idx].value

        yield record

    wb.close()


def extract_analytics_file(filepath):
    """提取达人属性表数据。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    # 第1行是表头
    headers = [_normalize_header(c.value) for c in sheet[1]]

    col_map = {
        "达人": "creator_name",
        "粉丝数": "followers_count",
        "量级": "tier",
        "博主人设": "persona",
        "账号属性/粉丝关注": "account_attr",
        "粉丝画像": "fan_portrait",
        "笔记画像": "notes_portrait",
        "投放CPE": "cpe",
        "近30天报备笔记数": "sponsored_notes",
        "爆文率": "viral_rate",
    }

    target_cols = {}
    for idx, h in enumerate(headers):
        if h in col_map:
            target_cols[idx] = col_map[h]

    records = []
    for row_idx in range(2, sheet.max_row + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue

        record = {"source_file": filepath.name, "source_row": row_idx}
        for col_idx, field_name in target_cols.items():
            record[field_name] = row[col_idx].value
        records.append(record)

    wb.close()
    return records
